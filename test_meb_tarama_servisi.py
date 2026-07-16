import unittest
from datetime import date
from io import BytesIO
from unittest.mock import MagicMock, patch

from docx import Document
from openpyxl import Workbook

import meb_tarama_servisi


class MebTaramaServisiTesti(unittest.TestCase):
    def test_ek_dosya_yeni_ilan_ve_alarm_sayilmaz(self):
        self.assertFalse(meb_tarama_servisi._yeni_ana_ilan_mi(1, "ek_dosya"))
        self.assertTrue(meb_tarama_servisi._yeni_ana_ilan_mi(1, "detay"))
        self.assertFalse(meb_tarama_servisi._yeni_ana_ilan_mi(0, "detay"))

    def test_tarihi_metin_icinden_duzeltir(self):
        self.assertEqual(
            "2026-07-14",
            meb_tarama_servisi._tarih_duzelt("Yayın tarihi: 14/07/2026 10:30"),
        )

    def test_bir_yillik_sinir_ve_tarihsiz_kayit(self):
        with patch.object(meb_tarama_servisi, "ihale_tarih_siniri", return_value=date(2025, 7, 15)):
            self.assertFalse(meb_tarama_servisi._tarih_gecerli("2025-07-14"))
            self.assertTrue(meb_tarama_servisi._tarih_gecerli("2025-07-15"))
            self.assertFalse(meb_tarama_servisi._tarih_gecerli(None))

    def test_ihale_tarihini_baglamindan_ayiklar(self):
        metin = """
            01.12.2023 tarihli yönetmelik hükümleri uygulanır.
            İhalenin Yapılacağı Tarih/Saat: 21.07.2026 Salı günü saat 10.00
        """
        self.assertEqual(
            "2026-07-21",
            meb_tarama_servisi._ihale_tarihi_bul(metin, "2026-07-14"),
        )
        self.assertEqual(
            "2026-07-31",
            meb_tarama_servisi._ihale_tarihi_bul(
                "b) Tarihi ve Saati : 31.07.2026 Cuma günü saat 15:00",
                "2026-07-13",
            ),
        )
        self.assertEqual(
            "2026-08-04",
            meb_tarama_servisi._ihale_tarihi_bul(
                "İHALE TARİHİ：04.08.2026 İHALE SAATİ:10:30",
                "2026-07-13",
            ),
        )

    def test_yayin_tarihini_ihale_tarihi_saymaz(self):
        self.assertIsNone(
            meb_tarama_servisi._ihale_tarihi_bul(
                "Yayın: 07.07.2026 Şehit Oğuz Özgür Çevik kantin ihale ilanı",
                "2026-07-07",
            )
        )
        self.assertEqual(
            "2026-07-20",
            meb_tarama_servisi._ihale_tarihi_bul(
                "Yayın: 07.07.2026 İhale Tarihi: 20.07.2026 İhale Saati: 14.00",
                "2026-07-07",
            ),
        )

    def test_kanun_numarasini_ihale_tarihi_saymaz(self):
        metin = """
            3.2.Tarihi ve saati
            4.2.5237 sayılı Türk Ceza Kanunu
            4.3.4734 sayılı Kanunun 11 inci maddesi
        """
        self.assertIsNone(
            meb_tarama_servisi._ihale_tarihi_bul(metin, "2025-10-03")
        )

    def test_bir_yildan_uzak_ihale_tarihini_reddeder(self):
        self.assertIsNone(
            meb_tarama_servisi._ihale_tarihi_bul(
                "İhale tarihi: 04.03.4734", "2025-10-03"
            )
        )

    def test_docx_tablosundan_ihale_tarihini_okur(self):
        belge = Document()
        tablo = belge.add_table(rows=1, cols=2)
        tablo.cell(0, 0).text = "İhale Tarihi"
        tablo.cell(0, 1).text = "20.07.2026 Pazartesi saat 14.00"
        icerik = BytesIO()
        belge.save(icerik)
        yanit = MagicMock(content=icerik.getvalue())
        yanit.raise_for_status.return_value = None
        oturum = MagicMock()
        oturum.get.return_value = yanit

        self.assertEqual(
            "2026-07-20",
            meb_tarama_servisi._docx_tarihi_getir(
                oturum, "https://mamak.meb.gov.tr/ihale.docx", "2026-07-07"
            ),
        )

    def test_xlsx_hucresinden_ihale_tarihini_okur(self):
        kitap = Workbook()
        sayfa = kitap.active
        sayfa.append(["İhale Tarihi", "20.07.2026 Pazartesi saat 14.00"])
        icerik = BytesIO()
        kitap.save(icerik)
        yanit = MagicMock(content=icerik.getvalue())
        yanit.raise_for_status.return_value = None
        oturum = MagicMock()
        oturum.get.return_value = yanit

        self.assertEqual(
            "2026-07-20",
            meb_tarama_servisi._xlsx_tarihi_getir(
                oturum, "https://mamak.meb.gov.tr/ihale.xlsx", "2026-07-07"
            ),
        )


if __name__ == "__main__":
    unittest.main()
